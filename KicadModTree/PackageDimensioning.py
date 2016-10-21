from openpyxl import load_workbook
import openpyxl
from collections import namedtuple, OrderedDict

SOPKG_TYPES = ['SOIC', 'SSOP', 'TSSOP', 'TVSOP', 'VSSOP']
SOT_TYPES = ['SOT23', 'SOT143', 'SOT223', 'SOFL']
SOD_TYPES = ['SOD', 'SODFL']

sopkg = namedtuple('sopkg', [
                   'name',     # Package name
                   'n',        # Number of pins
                   'D_min',    # Package width min
                   'D_max',    # Package width max
                   'E1_min',   # Package length min
                   'E1_max',   # Package length max
                   'E_min',    # Package length with leads min
                   'E_max',    # Package length with leads max
                   ### 'A1_max'
                   'A_max',    # Package height max
                   'L_min',    # Terminal length min
                   'L_max',    # Terminal length max
                   'b_min',    # Min lead width
                   'b_max',    # Max lead width
                   'e',])      # Pitch

sot = namedtuple('sot',[
                 'name',     # Name of component
                 'n',        # Number of pins
                 'D_min',    # Case width min
                 'D_max',    # Case width max
                 'E1_min',   # Case length min
                 'E1_max',   # Case length max
                 'E_min',    # Length with leads min
                 'E_max',    # Length with leads max
                 'A_max',    # Case height
                 'A1_max',   # Case to PCB distance
                 'L_min',    # Min lead length
                 'L_max',    # Max lead length
                 'b_min',    # Min lead width
                 'b_max',    # Max lead width
                 'b1_min',   # Possibly optional large lead
                 'b1_max',   # Possibly optional large lead
                 'e'])       # Pitch

sofl = namedtuple('sofl',[
                 'name',     # Name of component
                 'n',        # Number of pins
                 'D_min',    # Case width min
                 'D_max',    # Case width max
                 'E1_min',   # Case length min
                 'E1_max',   # Case length max
                 'E_min',    # Length with leads min
                 'E_max',    # Length with leads max
                 'A_max',    # Case height
                 'L_min',    # Min lead length
                 'L_max',    # Max lead length
                 'b_min',    # Min lead width
                 'b_max',    # Max lead width
                 'e'])       # Pitch

qfn = namedtuple('qfn', [
                 'name',
                 'n',
                 'D_min',
                 'D_max',
                 'E_min',
                 'E_max',
                 'A_max',
                 'L_min',
                 'L_max',
                 'b_min',
                 'b_max',
                 'e'])



def getPackageDims(paramtype, pkgtype):
    packages = OrderedDict()
    if paramtype == 'GullWingParams':
        wb = load_workbook('C:/KicadRepo/kicad-footprint-generator/KicadModTree/GullWingParams.xlsx')
        if pkgtype in SOPKG_TYPES:
            sheet = wb.get_sheet_by_name(pkgtype)
                ### TODO: Add e-pad dimensions
                ### TODO: Possibly add in via dimensions with the e-pads?
            for row in range(2, sheet.max_row + 1):
                # Each row in the spreadsheet has data for one footprint
                Name = sheet['A' + str(row)].value
                if Name == None:
                    break
                package = sopkg(
                    name = Name,
                    n = int(sheet['B' + str(row)].value),
                    D_min = sheet['C' + str(row)].value,
                    D_max = sheet['D' + str(row)].value,
                    E1_min = sheet['E' + str(row)].value,
                    E1_max = sheet['F' + str(row)].value,
                    E_min = sheet['G' + str(row)].value,
                    E_max = sheet['H' + str(row)].value,
                    A_max = sheet['I' + str(row)].value,
                    L_min = sheet['J' + str(row)].value,
                    L_max = sheet['K' + str(row)].value,
                    b_min = sheet['L' + str(row)].value,
                    b_max = sheet['M' + str(row)].value,
                    e = sheet['N' + str(row)].value)
                packages[Name] = package
        if pkgtype in SOT_TYPES:
            sheet = wb.get_sheet_by_name(pkgtype)
            
    return packages

    if paramtype == 'PthParams':
        wb = load_workbook('C:/KicadRepo/kicad-footprint-generator/KicadModTree/PthParams.xlsx')
        if pkgtype == 'Resistor':
        elif pkgtype == 'Capacitor':
        elif pkgtype == 'Inductor':