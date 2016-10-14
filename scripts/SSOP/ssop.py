#!/usr/bin/env python

import sys
import os
from collections import namedtuple

sys.path.append("..\\..")
from KicadModTree import * # NOQA
from KicadModTree.nodes.specialized.PadArray import PadArray
from KicadModTree.nodes.base.Pad import Pad
import KicadModTree.Rules as rules
from openpyxl import load_workbook
from KicadModTree.Tolerancing import GullWingParams

print('Opening workbook...')
wb = openpyxl.load_workbook()
sheet = wb.get_sheet_by_name('Population by Census Tract')
data = {}
# Fill in data with each footprints D parameters
print('Reading rows...')

def getFootprintMeasurements():
    for row in range(2, sheet.max_row + 1):
        # Each row in the spreadsheet has data for one footprint
        Name = sheet['A' + str(row)].value
        n = sheet['B' + str(row)].value
        D_min = sheet['C' + str(row)].value
        D_max = sheet['D' + str(row)].value
        data[Name] = {n, D_min, D_max}


DEFAULT_DENSITY_LEVEL = 'B' # Nominal

prefix = "SSOP-"

desc = "SSOP package, {pnum}"
# ssop lead dimensions; used for fabrication outlines of pads
# length, width; 0.05mm line width fab layer; measurements are maximums
ssop_lead_dims = [0.95, 0.38]

for k in ssop.keys():
    
    prop = ssop[k]

    name = prefix + k + 

    layers = ["F.Cu", "F.Paste", "F.Mask"]

    footprint_name = prefix + pins + dims

    kicad_mod = Footprint(footprint_name)
    kicad_mod.setDescription("test")
    kicad_mod.setAttribute("smd")
    kicad_mod.setTags("test")

    kicad_mod.append(Text(type='reference', text='REF**', at=[0, -3], layer='F.SilkS'))
    kicad_mod.append(Text(type='value', text=footprint_name, at=[1.5, 3], layer='F.Fab'))

    
    # **TODO** Implement zero orientation function
    # if ZERO_ORIENTATION == 'LevelA': # pin 1 in upper left corner
    #     l,w = w,l
    # else 'LevelB': # pin 1 in lower left corner; rotated 90 from LevelA
    #     continue
    pin_y_pos = (bl_lmax-b_lmax)/2+b_lmax
    pin_y_neg = -(bl_lmax-b_lmax)/2-b_lmax
    pad_l_start = [-(pitch*pnum/4-pitch/2), pin_y_pos/2]
    pad_r_start = [-(pitch*pnum/4-pitch/2), -pin_y_pos/2]

    pad_num_r_start = pnum/2 + 1
    # draw the pads
    pa_left = PadArray(pincount=pnum/2, intial=1, start=pad_l_start, increment=1, layers=layers, shape=Pad.SHAPE_ROUNDRECT, roundrect_ratio=0.25, type = Pad.TYPE_SMT, size = [0.4, 1.2], x_spacing = pitch)
    pa_right = PadArray(pincount=pnum/2, initial=pnum, start=pad_r_start, increment=-1, layers=layers, shape=Pad.SHAPE_ROUNDRECT, roundrect_ratio=0.25, type=Pad.TYPE_SMT, size=[0.4, 1.2], x_spacing = pitch)
    kicad_mod.append(pa_left)
    kicad_mod.append(pa_right)

    # create silkscreen
    # kicad_mod.append()

    # add designator for pin #1

    # draw the courtyard

    # kicad_mod.append(RectLine(start=[], end=[], layer='F.CrtYd'))

    # draw the fabrication outline
    kicad_mod.append(RectLine(start=[b_wmax/2, -b_lmax/2], end=[-b_wmax/2, b_lmax/2], layer='F.Fab', width=0.15))

    # draw fab outlines for pads if selected as option
    # **TODO** if DRAW_PAD_FAB:
        
    # **TODO** Have it get the zero orientation first

    # add model
    kicad_mod.append(Model(filename="Housings_SSOP/{fp_name}.{model_type}".format(fp_name = footprint_name, model_type = rules.MODEL_3D_TYPE),
                           at=[0, 0, 0], scale=[1, 1, 1], rotate=[0, 0, 0]))

    print(kicad_mod.getCompleteRenderTree())

    filename = prefix + "{pins}{l}x{w}mm_{p}mm".format(pins=pins, l=nominal_l, w=nominal_w, p=pitch) + ".kicad_mod"
    # write file
    file_handler = KicadFileHandler(kicad_mod)
    file_handler.writeFile(filename)